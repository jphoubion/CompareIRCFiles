import sys
from pprint import pprint
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter


from PySide6 import QtWidgets, QtCore
from PySide6.QtCore import Qt

from PySide6.QtWidgets import QDialog, QFileDialog, QLabel, QMessageBox

from ui.MainWindow import Ui_MainWindow

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self.btn_old_file.clicked.connect(self.select_old_file)
        self.btn_new_file.clicked.connect(self.select_new_file)
        self.btn_Quit.clicked.connect(quit)
        self.btn_compare.clicked.connect(self.compare_table)
        self.btn_export.clicked.connect(self.export_table)

        self.old_filename = ""
        self.new_filename = ""
        self.old_content = []
        self.new_content = []
        self.result_content = []
        self.cpt_new = 0
        self.cpt_chg = 0
        self.nb_result = 0

        self.workbook = None
        self.sheet = None
        self.line_to_color = []


    def select_old_file(self, s):
        dlg = QFileDialog(self, filter="*.csv")
        dlg.setWindowTitle("Selectionner le fichier actuel")
        dlg.setFileMode(QFileDialog.ExistingFile)
        dlg.setViewMode(QFileDialog.Detail)
        if dlg.exec():
            self.old_filename = dlg.selectedFiles()
            self.lbl_old_filename.setText(self.old_filename[0])
            # print(f"old_filename = {self.old_filename}")
            with open(self.old_filename[0], "r") as old_file:
                self.old_content = old_file.readlines()
            self.fill_old_file_tv(self.old_content)

    def fill_old_file_tv(self, old_content):
        self.tw_old_file.setRowCount(len(old_content))
        nb_col = len(old_content[0].split(';'))
        self.tw_old_file.setColumnCount(nb_col)
        self.tw_old_file.setAlternatingRowColors(True)
        # print(old_content[0])
        self.tw_old_file.setHorizontalHeaderLabels(old_content[0].split(';'))
        for content in range(1, len(old_content)):
            # print(old_content[content])
            for field in range(nb_col):
                # print(old_content[content].split(";")[field])
                lbl = QLabel(old_content[content].split(";")[field])
                self.tw_old_file.setCellWidget(content-1, field, lbl)

        self.tw_old_file.resizeRowsToContents()
        self.tw_old_file.resizeColumnsToContents()


    def select_new_file(self):
        dlg = QFileDialog(self, filter="*.csv")
        dlg.setWindowTitle("Selectionner le nouveau fichier")
        dlg.setFileMode(QFileDialog.ExistingFile)
        dlg.setViewMode(QFileDialog.Detail)
        if dlg.exec():
            self.new_filename = dlg.selectedFiles()
            self.lbl_new_filename.setText(self.new_filename[0])
            # print(f"new_filename = {self.new_filename}")
            with open(self.new_filename[0], "r") as new_file:
                self.new_content = new_file.readlines()
            self.fill_new_file_tv(self.new_content)

    def fill_new_file_tv(self, new_content):
        self.tw_new_file.setRowCount(len(new_content))
        nb_col = len(new_content[0].split(';'))
        self.tw_new_file.setColumnCount(nb_col)
        self.tw_new_file.setAlternatingRowColors(True)
        # print(new_content[0])
        self.tw_new_file.setHorizontalHeaderLabels(new_content[0].split(';'))
        for content in range(1, len(new_content)):
            # print(new_content[content])
            for field in range(nb_col):
                # print(new_content[content].split(";")[field])
                lbl = QLabel(new_content[content].split(";")[field])
                self.tw_new_file.setCellWidget(content-1, field, lbl)

        self.tw_new_file.resizeRowsToContents()
        self.tw_new_file.resizeColumnsToContents()

    def compare_table(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active

        cpt = 0

        # initialisation de la table
        # self.tw_result.row
        self.tw_result.setRowCount(len(self.old_content))
        nb_col = len(self.old_content[0].split(';'))
        self.tw_result.setColumnCount(nb_col)
        self.tw_result.setAlternatingRowColors(True)
        self.tw_result.setHorizontalHeaderLabels(self.new_content[0].split(';'))

        # boucle sur le NOUVEAU fichier
        for new_c in range(len(self.new_content)):
            # print(new_c)
            found = False
            # Boucle sur le fichier ACTUEL
            for old_c in range(len(self.old_content)):
                if self.new_content[new_c].split(";")[1] == self.old_content[old_c].split(";")[1]: # check sur le numéro dossier IRC
                    found = True
                    # print("même client >>>>>>>>>>>>>>>>>>")
                    # print(self.new_content[new_c].split(";")[1])
                    # print(self.old_content[old_c].split(";")[1])
                    # found = True
                    if self.new_content[new_c].split(";")[3] != self.old_content[old_c].split(";")[3]:
                        # print("contenu different")
                        # print(self.new_content[new_c].split(";")[3])
                        # print(self.old_content[old_c].split(";")[3])
                        # print(">"*50)
                        self.cpt_chg += 1

                        # rowPosition = self.tw_result.rowCount()
                        # self.tw_result.insertRow(rowPosition)

                        self.result_content.append(self.old_content[old_c].split(";"))
                        self.result_content.append(self.new_content[new_c].split(";"))

                        # Ajout dans la sheet XLS si on veut l'exporter
                        self.sheet.append(self.old_content[old_c].split(";"))
                        self.sheet.append(self.new_content[new_c].split(";"))

                        # ajout dans la table de resultat
                        self.nb_result += 2
                        for field in range(nb_col):
                            # Ajout de la version actuelle
                            lbl = QLabel(self.old_content[old_c].split(";")[field])
                            self.tw_result.setCellWidget(self.nb_result-1, field, lbl)
                            lbl.setContentsMargins(5, 5, 5, 5)
                            lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter)

                            # Ajout de la mise à jour
                            lbl2 = QLabel(self.new_content[new_c].split(";")[field])
                            self.tw_result.setCellWidget(self.nb_result, field, lbl2)
                            lbl2.setContentsMargins(5, 5, 5, 5)
                            lbl2.setAlignment(Qt.AlignmentFlag.AlignVCenter)
                    break

            if found == False:
                # print("NOUVEAU")
                # print(self.new_content[new_c].split(";")[1])
                # print(self.old_content[old_c].split(";")[1])
                self.cpt_new += 1
                self.nb_result += 1

                for field in range(nb_col):
                    # print(new_content[content].split(";")[field])
                    lbl = QLabel(self.new_content[new_c].split(";")[field])
                    lbl.setStyleSheet("background-color:green")
                    self.tw_result.setCellWidget(self.nb_result, field, lbl)
                    lbl.setContentsMargins(5, 5, 5, 5)
                    lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter)
                self.result_content.append(self.new_content[new_c].split(";"))
                # Ajout des nouvelles lignes dans le sheet XLS
                self.sheet.append(self.new_content[new_c].split(";"))
                self.line_to_color.append(f"{self.nb_result}")

                # break
        self.tw_result.resizeRowsToContents()
        self.tw_result.resizeColumnsToContents()

    def export_table(self):
        #Coloration des lignes
        for line in self.line_to_color:
            for c in range(1,12):
                col = get_column_letter(c)
                self.sheet[f"{col}{line}"].fill = PatternFill("solid", fgColor="00FF00")

        # Choose filename
        dialog_save = QFileDialog()
        name = dialog_save.getSaveFileName(filter="*.xlsx;; *.xls")
        if name[0] !="":
            self.workbook.save(name[0])
            dialog_info = QMessageBox(self)
            dialog_info.setWindowTitle("Information")
            dialog_info.setText("Fichier Excel exporté !")
            dialog_info.setStandardButtons(QMessageBox.Yes)
            dialog_info.setIcon(QMessageBox.Information)
            btn = dialog_info.exec()

def quit(self):
    app.quit()


app = QtWidgets.QApplication(sys.argv)

window = MainWindow()
window.show()
app.exec()